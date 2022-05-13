from datetime import datetime

import openpyxl
import requests
from googletrans import Translator

from config import open_weather_token
from news import news


def excell(data):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = "Місто"
    sheet['B1'] = "Температура"
    sheet['C1'] = "Погода"
    sheet['D1'] = "Вологість"
    sheet['E1'] = "Хмарність"
    sheet['G1'] = "Звіт зроблений"
    sheet['F1'] = "Видимість"
    sheet['H1'] = "Відчувається як"
    sheet['I1'] = "Максимальна температура повітня"
    sheet['J1'] = "Мінімальна температура повітря"
    sheet['K1'] = "Тиск"
    sheet['L1'] = "Опис погоди"
    sheet['M1'] = "Швидкість вітру"
    sheet['N1'] = "Напрям вітру(градусів)"
    sheet['O1'] = "Поривність вітру"

    row = 2

    sheet[row][0].value = f'{data["name"]}'
    sheet[row][1].value = f'{data["main"]["temp"]} C°'
    sheet[row][2].value = data["weather"][0]["main"]
    sheet[row][3].value = f'{data["main"]["humidity"]} %'
    sheet[row][4].value = f'{data["clouds"]["all"]} %'
    sheet[row][5].value = f'{data["visibility"]} м'
    sheet[row][6].value = datetime.now().strftime('%Y-%m-%d %H:%M')
    sheet[row][7].value = f'{data["main"]["feels_like"]} C°'
    sheet[row][8].value = f'{data["main"]["temp_max"]} C°'
    sheet[row][9].value = f'{data["main"]["temp_min"]} C°'
    sheet[row][10].value = f'{data["main"]["pressure"]} мм.рт.ст'
    sheet[row][11].value = f'{data["weather"][0]["description"]}'
    sheet[row][12].value = f'{data["wind"]["speed"]} м/с'
    sheet[row][13].value = f'{data["wind"]["deg"]}°'
    sheet[row][14].value = f'{data["wind"]["gust"]} м/с'
    row += 1

    book.save("Report.xlsx")
    book.close()


def excell_more_days(data, days=1):
    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = "Місто"
    sheet['B1'] = "Температура"
    sheet['C1'] = "Погода"
    sheet['D1'] = "Вологість"
    sheet['E1'] = "Хмарність"
    sheet['G1'] = "Звіт зроблений"
    sheet['F1'] = "Видимість"
    sheet['H1'] = "Відчувається як"
    sheet['I1'] = "Максимальна температура повітня"
    sheet['J1'] = "Мінімальна температура повітря"
    sheet['K1'] = "Тиск"
    sheet['L1'] = "Опис погоди"
    sheet['M1'] = "Швидкість вітру"
    sheet['N1'] = "Напрям вітру(градусів)"
    sheet['O1'] = "Поривність вітру"

    row = 2
    i = 0

    while row <= int(days) * 8:
        sheet[row][0].value = f'{data["city"]["name"]}'
        sheet[row][1].value = f'{data["list"][i]["main"]["temp"]} C°'
        sheet[row][2].value = data["list"][i]["weather"][0]["main"]
        sheet[row][3].value = f'{data["list"][i]["main"]["humidity"]} %'
        sheet[row][4].value = f'{data["list"][i]["clouds"]["all"]} %'
        sheet[row][5].value = f'{data["list"][i]["visibility"]} м'
        sheet[row][6].value = data['list'][i]['dt_txt']
        sheet[row][7].value = f'{data["list"][i]["main"]["feels_like"]} C°'
        sheet[row][8].value = f'{data["list"][i]["main"]["temp_max"]} C°'
        sheet[row][9].value = f'{data["list"][i]["main"]["temp_min"]} C°'
        sheet[row][10].value = f'{data["list"][i]["main"]["pressure"]} мм.рт.ст'
        sheet[row][11].value = f'{data["list"][i]["weather"][0]["description"]}'
        sheet[row][12].value = f'{data["list"][i]["wind"]["speed"]} м/с'
        sheet[row][13].value = f'{data["list"][i]["wind"]["deg"]}°'
        sheet[row][14].value = f'{data["list"][i]["wind"]["gust"]} м/с'
        row += 1
        i += 1

    book.save("Report.xlsx")
    book.close()

def excell_news(data,news):
    translator = Translator()
    book = openpyxl.Workbook()
    sheet = book.active
    sheet['A1'] = "Запит"
    sheet['B1'] = "Заголовок"
    sheet['C1'] = "Контент"
    sheet['D1'] = "Опис"
    sheet['E1'] = "Автор"
    sheet['F1'] = "Джерело"
    sheet['G1'] = "URL статті"
    sheet['H1'] = "URL фото"
    sheet['I1']="Публікація статті"
    row = 2
    i=0
    datas=data["data"]
    lens=len(datas["articles"])
    while i<lens:
        sheet[row][0].value = f'{translator.translate((news), dest="uk").text}'
        sheet[row][1].value = f'{translator.translate((datas["articles"][i]["title"]), dest="uk").text}°'
        sheet[row][2].value = f'{translator.translate((datas["articles"][i]["content"]), dest="uk").text}'
        sheet[row][3].value = f'{translator.translate((datas["articles"][i]["description"]), dest="uk").text} '
        sheet[row][4].value = f'{translator.translate((datas["articles"][i]["author"]), dest="uk").text}'
        sheet[row][5].value = f'{datas["articles"][i]["source"]["name"]}'
        sheet[row][6].value = f'{datas["articles"][i]["url"]}'
        sheet[row][7].value = f'{datas["articles"][i]["urlToImage"]}'
        sheet[row][8].value = f'{datas["articles"][i]["publishedAt"]}'
        i+=1
        row += 1
    book.save("Report_news.xlsx")
    book.close()